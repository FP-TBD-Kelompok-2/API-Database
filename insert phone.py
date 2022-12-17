import requests
import openpyxl

path = "phone_dataset.xlsx"

wb_obj = openpyxl.load_workbook(path)
worksheet = wb_obj.active


ColNames = {}
Current = 0
for COL in worksheet.iter_cols(1, worksheet.max_column):
    ColNames[COL[0].value] = Current
    Current += 1

for row_cells in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
    if row_cells[ColNames['device_name']].value is not None and row_cells[ColNames['input_flag']].value != 1:

        row_cells[ColNames['input_flag']].value = 1

        device_name = row_cells[ColNames['device_name']].value
        device_brand = row_cells[ColNames['device_brand']].value
        device_storage = row_cells[ColNames['device_storage']].value
        device_connectivity = row_cells[ColNames['device_connectivity']].value
        device_battery = row_cells[ColNames['device_battery']].value
        cpu_name = row_cells[ColNames['cpu_name']].value
        cpu_cores = row_cells[ColNames['cpu_cores']].value
        cpu_clock = row_cells[ColNames['cpu_clock']].value
        antutu_score = row_cells[ColNames['antutu_score']].value
        description = row_cells[ColNames['description']].value
        stock = row_cells[ColNames['stock']].value
        price = row_cells[ColNames['price']].value
        color = row_cells[ColNames['color']].value
        img = row_cells[ColNames['img']].value

        body = {
            "device_name": device_name,
            "device_brand": device_brand,
            "device_storage": device_storage,
            "device_connectivity": device_connectivity,
            "device_battery": device_battery,
            "cpu_name": cpu_name,
            "cpu_cores": cpu_cores,
            "cpu_clock": cpu_clock,
            "antutu_score": antutu_score,
            "description": description,
            "stock": stock,
            "price": price,
            "color": color,
            "img" : img
        }
        x = requests.post('https://formal-audio-370910.as.r.appspot.com/hp/add', json=body)
        print(x.text)
wb_obj.save("phone_dataset.xlsx")